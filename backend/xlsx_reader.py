"""
Чтение локального .xlsx (выгрузка той же таблицы) в формат (rows, formulas, tz),
совместимый с parser.build_model и migrate-логикой.

Нужен для офлайн-миграции и тестов: даёт ту же прямоугольную сетку значений и формул,
что и sheet.fetch_grid() для живой Google-таблицы, но из файла.

  rows[r][c]     — datetime.date для дат, float/str/bool или None;
  formulas[r][c] — строка-формула ('' если ячейка не формула).

openpyxl читаем дважды: data_only=True даёт вычисленные значения, data_only=False —
тексты формул (нужны для определения диапазона колонок строки «Осталось»).
"""

from datetime import datetime

import openpyxl


def _value(cell):
    v = cell.value
    if isinstance(v, datetime):
        return v.date()
    return v


def read_grid(path, sheet_name=None, tz='Europe/Moscow'):
    """Возвращает (rows, formulas, tz) — прямоугольные списки списков."""
    wb_v = openpyxl.load_workbook(path, data_only=True, read_only=True)
    wb_f = openpyxl.load_workbook(path, data_only=False, read_only=True)
    ws_v = wb_v[sheet_name] if sheet_name else wb_v.worksheets[0]
    ws_f = wb_f[sheet_name] if sheet_name else wb_f.worksheets[0]

    rows, formulas = [], []
    width = 0
    for row_v, row_f in zip(ws_v.iter_rows(), ws_f.iter_rows()):
        vrow = [_value(c) for c in row_v]
        frow = []
        for c in row_f:
            fv = c.value
            frow.append(str(fv) if isinstance(fv, str) and fv.startswith('=') else '')
        width = max(width, len(vrow), len(frow))
        rows.append(vrow)
        formulas.append(frow)

    wb_v.close()
    wb_f.close()

    for vrow, frow in zip(rows, formulas):
        if len(vrow) < width:
            vrow.extend([None] * (width - len(vrow)))
        if len(frow) < width:
            frow.extend([''] * (width - len(frow)))
    return rows, formulas, tz
